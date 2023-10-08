import openpyxl as xl
from openpyxl.chart import PieChart, Reference


def process_workbook(excel_file):
    workbook = xl.load_workbook(excel_file)
    sheet = workbook['Sheet1']

    for row in range(2, sheet.max_row + 1):
        # Let's reach for all the rows in column 3 (which is the price)
        cell = sheet.cell(row, 3)
        # then multiply all the values in row 2 , 3, 4; col 3 by 0.9
        corrected_price = cell.value * 0.9

        # Let's add the corrected_price in a new column
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    value = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)
    # add the data from the value into the pie chart
    chart = PieChart()
    chart.add_data(value)
    # add the pie chart in the sheet (specify the chart and the row and col
    sheet.add_chart(chart, 'e2')

    workbook.save(excel_file)